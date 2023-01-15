# openpyxl是处理excel文件的python第三方库
# 安装: pip3 install openpyxl
# 描述读取excel工作簿、工作表、行列、单元格属性的代码演示


import openpyxl

wb = openpyxl.load_workbook('test.xlsx')



# 打印表单名称
for sheet in wb:
    print(sheet.title)

# 在wb工作簿中新建表单
mySheet = wb.create_sheet("赵慧芳")

print(wb.sheetnames)


s1 = wb.get_sheet_by_name('Sheet1')
print(s1.title)


sActive = wb.active
print(sActive.title)
print(sActive)
print(sActive['A1'])
print(sActive['A1'].value)

c = sActive['A1']

print('row is {}, col is {}, val is {}'.format(c.row, c.column, c.value))

print('cell {}, val is {}'.format(c.coordinate,c.value))

print(sActive.cell(row=1,column=1).value)

for i in range(2, 8, 2):
    print(i, sActive.cell(row=i, column=1).value)


cCol = sActive['A']
print(cCol)


col_range = sActive['B:C']
row_range = sActive['2:6']

for col in col_range:
    for cell in col:
        print(cell.value)


for row in row_range:
    for cell in row:
        print(cell.value)

print('\n')

for row in sActive.iter_rows(min_row=2,max_row=6,min_col=2,max_col=3):
    for cel in row:
        print(cel.value)

print('\n')
print('\n')

print(tuple(sActive.rows))


cell_range = sActive['A1:B3']


print('\n')
print('\n')
print('\n')
for rowCell in cell_range:
    for cell in rowCell:
        print(cell.coordinate, cell.value)
    print('------end of row ------ ')


print('{} * {}'.format(sActive.max_row, sActive.max_column))
